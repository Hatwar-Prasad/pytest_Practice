import openpyxl
import pytest
from selenium.webdriver.common.by import By
test = openpyxl.load_workbook('C:/Users/drpra/PycharmProjects/pytest_Practice/testData/ReadingData.xlsx')
data = test['testData']


# def datagenerator():
#
#     data_rows = data.max_row
#     data_cols = data.max_column
#
#     for i in range(1, data_rows):
#         # List1 = []
#         for j in range(1, data_cols + 1):
#             cell_value = data.cell(i, j)
#             # List1.append(cell_value.value)



#
A=data.cell(row=2,column=1).value
B=data.cell(row=2,column=2).value
C=data.cell(row=3,column=1).value
D=data.cell(row=3,column=2).value

for row in range(data.max_row):
    for col in data.iter_cols( data.max_column):

        b = col[row].value

class Test_Login:

    # @pytest.mark.parametrize("username,password",
    #                      [("raj", "raj1234"), ("prasad", "prasad1234"), ("rohit", "rohit1234")])  # parameterized

    @pytest.mark.parametrize("username,password",[(A, B), (C, D)])
    def test_login(self, setup_and_teardown, username,password):
        print("this is login test case")

        # expectedTitle="Facebook – log in or sign up"   #assertion
        # assert expectedTitle==driver.title
        self.driver.find_element(By.ID, "email").send_keys(username)
        self.driver.find_element(By.ID, "pass").send_keys(password)
